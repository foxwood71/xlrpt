#!/usr/bin/python
# -*- coding: utf-8 -*-

"""
    cimon일보를 가져와 월보와 년보를 생성하는 프로그램.
    ' 오포관련 특기사항 반영 및 프로그램 구성변경
    Usage:
        # Command line run
        # C:>start /b python xlrpt.py -p ga -s 2023-02-01 -e 2023-02-28 -t m
"""

import datetime
import os
import sys
import time

import openpyxl
import win32com.client
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string, coordinate_to_tuple

import xlrpt_utils

FLOW = "f"  # 유량

DAILY = "d"  # 일보
MONTHLY = "m"  # 월보
YEARLY = "y"  # 년보


DAILY_RPT_DATA_HEIGHT = 24
MONTHLY_RPT_DATA_HEIGHT = 31
YEARLY_RPT_DATA_HEIGHT = 12


# rc address to 'A1' address
def xlref(row, column, zero_indexed=True):
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)


# def date_range(start_date, end_date):
#     for n in range(int ((end_date - start_date).days+1)):
#         yield start_date + datetime.timedelta(n)


def range_value(sheet, rng_x, rng_height, rng_y, rng_width):
    """
    range_value
    """
    return tuple(
        tuple(sheet.cell(row=rng_x + i, column=rng_y + j).value for j in range(rng_width - rng_y + 1))
        for i in range(rng_width - rng_x + 1)
    )


def xlsx_rpt(**kwargs):
    """
    xlsx_rpt 함수 동작
        키워드 인수를 받아 엑셀 보고서를 생성하는 함수
    키보드 인수
        conf = 보고서 설정(JSON 형식)
        stp : 보고서 작성대상 시설명 str
        rpt_type : 보고서 종류 op, dbf, ...
        rpt_cyc : 보고서 주기 m, y
        start_date : 보고서 작성 시작일자
        end_date : 보고서 작성 종료일자
    """
    conf: dict = kwargs["conf"]
    stp: str = kwargs["stp"]
    rpt_type: str = kwargs["rpt_type"]

    src_rpt_cycl: str = ""
    dst_rpt_cycl: str = kwargs["rpt_cycl"]

    start_date: datetime.date = kwargs["start_date"]
    end_date: datetime.date = kwargs["end_date"]

    # [ 표현식 for 항목 in 리스트 or 튜플 if 조건문 ]
    # 변수 힌팅 및 초기화 ㅜㅡㅜ 조건문 내부 사용시 미초기화 변수로 인식
    rpt_range: range = range(0, 1, 1)

    rpt_type_root: str = f"{conf['report_root_path']}/{conf[stp]['folder']}/{conf[stp][rpt_type]['folder']}"

    src_wb_path: str = ""
    src_wb_name_prefix: str = ""
    dst_wb_path: str = ""
    dst_wb_name: str = ""
    dst_wb_template: str = ""

    #   wb_sum_cell_types: list[list[list[int]]] = [[[]]]
    #   wb_shts: int = 0
    wb_pgs_in_shts: list[int] = []
    wb_cols_in_shts: list[list[int]] = []

    pg_size_src: int = 0
    pg_size_dst: int = 0

    first_data_rd_cell_in_shts: list[str] = []
    date_cell_in_shts: list[str] = []
    first_data_cell_in_shts: list[str] = []

    # 보고서 공용 변수 설정
    wb_sum_cell_types: dict = conf[stp][rpt_type]["sum_cell_types"]
    wb_shts: int = len(wb_sum_cell_types)

    src_rpt_date_height: int = 0
    dst_rpt_date_height: int = 0

    for i in range(wb_shts):
        pages = len(wb_sum_cell_types[i])
        wb_pgs_in_shts.append(pages)
        columns = []
        for j in range(pages):
            columns.append(len(wb_sum_cell_types[i][j]))
        wb_cols_in_shts.append(columns)

    date_cell_in_shts = conf[stp][rpt_type][dst_rpt_cycl]["date_cell_in_shts"]
    first_data_cell_in_shts = conf[stp][rpt_type][dst_rpt_cycl]["first_data_cell_in_shts"]

    if dst_rpt_cycl == MONTHLY:  # MONTHLY
        src_rpt_cycl = DAILY
        src_rpt_date_height = DAILY_RPT_DATA_HEIGHT
        dst_rpt_date_height = MONTHLY_RPT_DATA_HEIGHT

        rpt_range = range(start_date.day, end_date.day + 1, 1)

        src_wb_path = f"{rpt_type_root}/{start_date.year:04}/{start_date.month:02}"
        src_wb_name_prefix = f"{start_date.year:04}{start_date.month:02}"

        dst_wb_template = f"{conf['template_path']}/{conf[stp][rpt_type][dst_rpt_cycl]['template']}"
        dst_wb_path = f"{rpt_type_root}/{start_date.year:04}"
        dst_wb_name = f"{start_date.year:04}{start_date.month:02}.xlsx"

        first_data_rd_cell_in_shts = conf[stp][rpt_type][src_rpt_cycl]["first_sum_cell_in_shts"]

    elif dst_rpt_cycl == YEARLY:
        src_rpt_cycl = MONTHLY
        src_rpt_date_height = MONTHLY_RPT_DATA_HEIGHT
        dst_rpt_date_height = YEARLY_RPT_DATA_HEIGHT

        rpt_range = range(1, 13, 1)  # 년단위 보고

        src_wb_path = f"{rpt_type_root}/{start_date.year:04}"
        src_wb_name_prefix = f"{start_date.year:04}"

        dst_wb_template = f"{conf['template_path']}/{conf[stp][rpt_type][dst_rpt_cycl]['template']}"
        dst_wb_path = f"{rpt_type_root}"
        dst_wb_name = f"{start_date.year:04}.xlsx"

        first_data_rd_cell_in_shts = conf[stp][rpt_type][src_rpt_cycl]["first_sum_cell_in_shts"]

        # region
        """
        elif rpt_type == FLOW:  # 경안 특수보고서
        rpt_range = range(start_date.day, end_date.day + 1, 1)
        rpt_folder = f'{conf["report_root_path"]}/{conf[stp]["code"]}/{start_date.year:04}/{start_date.month:02}/'
        src_wb_name_prefix = f"{rpt_folder}{start_date.year:04}{start_date.month:02}"
        # => postfix str 검토
        wb_dest_name = (
            f'{conf["report_root_path"]}/{conf[stp]["code"]}/{start_date.year:04}/'
            f"{start_date.year:04}{start_date.month:02}_flow.xlsx"
        )

        wb_dst_template = conf[stp]["flow"]["template"]

                # pg_size_src = (
                #     conf[stp]['flow']['pg_head_size']
                #     + DAILY_RPT_DATA_HEIGHT
                #     + conf[stp]['flow']['pg_tail_size']
                # )
                # pg_size_dst = (
                #     conf[stp]['flow']['pg_head_size']
                #     + MONTHLY_RPT_DATA_HEIGHT
                #     + conf[stp]['flow']['pg_tail_size']
                # )

        first_data_rd_cell_in_shts = conf[stp]["daily"]["first_data_cell_in_shts"]
        date_cell_in_shts = conf[stp]["flow"]["date_cell_in_shts"]
        first_data_cell_in_shts = conf[stp]["flow"]["first_data_cell_in_shts"]
        """
    # endregion

    else:
        pass

    # region
    # 모든 시트를 반복하며 병합된 셀의 범위를 읽어와 리스트 형태로 mcr_coord_list에 추가(보류)
    """
        for sht_name in dst_wb.sheetnames:
            sht = dst_wb[sht_name]
            mcr_coord_list = [mcr.coord for mcr in sht.merged_cells.ranges]
    """
    # endregion

    # open template file
    dst_wb = openpyxl.load_workbook(dst_wb_template)

    # daily monthly yearly report common area
    # read_only=True, data_only=True -> cell value not return formulae
    for rpt_idx in rpt_range:
        try:
            src_wb = openpyxl.load_workbook(
                f"{src_wb_path}/{src_wb_name_prefix}{str(rpt_idx).zfill(2)}.xlsx",
                read_only=True,
                data_only=True,
            )
        except FileNotFoundError:
            print(f"Not found => {src_wb_path}/{src_wb_name_prefix}{str(rpt_idx).zfill(2)}.xlsx")
            continue
        except PermissionError:
            print("Please close excel file and restart program")
            sys.exit(1)
        # except Exception as e:
        #     raise
        else:  # no except run
            print(f"processing => {src_wb_path}/{src_wb_name_prefix}{str(rpt_idx).zfill(2)}.xlsx")

            pg_offset = 0

            if rpt_type == FLOW:  # ga-stp special report
                # region[ flow report
                """
                first_data_rd_cell_in_shts_a1 = coordinate_from_string(first_data_rd_cell_in_shts)

                first_data_cell_in_sht_rc = coordinate_from_string(first_data_cell_in_shts)

                for flow_idx, flow_ws_name in enumerate(conf[stp]["flow"]["shts"]):
                    ws_sht_conf = conf[stp]["flow"]["shts"][flow_ws_name]
                    ws_flow_sht = ws_sht_conf["sht"]
                    ws_flow_pg = ws_sht_conf["pg"]
                    ws_flow_col = ws_sht_conf["col"]

                    pg_flow_first_cell = ws_flow_col + str(
                        first_data_rd_cell_in_shts_a1[1] + ((pg_size_src * ws_flow_pg) + pg_offset)
                    )
                    pg_flow_last_cell = ws_flow_col + str(
                        first_data_rd_cell_in_shts_a1[1]
                        + str(DAILY_RPT_DATA_HEIGHT - 1)
                        + str((pg_size_src * ws_flow_pg) + pg_offset)
                    )

                    flow_rng = f"{pg_flow_first_cell}:{pg_flow_last_cell}"

                    src_ws = src_wb.worksheets[ws_flow_sht]
                    xl_flow_rng = src_ws[flow_rng]

                    dst_ws = dst_wb.worksheets[flow_idx]

                    # Date Print
                    if rpt_idx == 1:
                        date_cell_in_shts_rc = coordinate_from_string(date_cell_in_shts)
                        pg_date_cell = (
                            f"{date_cell_in_shts_rc[0]}"
                            f"{str(date_cell_in_shts_rc[1] + ((pg_size_dst * ws_flow_pg) + pg_offset))}"
                        )
                        dst_ws[pg_date_cell] = datetime.datetime.strptime(
                            f"{start_date.year:04}-{start_date.month:02}-01", "%Y-%m-%d"
                        )

                    # Data Print
                    for i in range(0, DAILY_RPT_DATA_HEIGHT, 1):
                        r1 = first_data_cell_in_sht_rc[1] + ((pg_size_dst * ws_flow_pg) + rpt_idx + pg_offset) - 1
                        c1 = column_index_from_string(first_data_cell_in_sht_rc[0]) + i
                        dst_ws.cell(row=r1, column=c1).value = xl_flow_rng[i][0].value
                """
            # endregion
            else:  # monthly yealy report
                # monthly yearly report sheet level
                for sht in range(0, wb_shts, 1):
                    # region - sheet processing print for debugging
                    """if (sht == wb_shts -1):
                        print(',' + str(sht))
                    else:
                        print(','+ str(sht), end=' ')"""
                    # endregion

                    # sht
                    src_ws = src_wb.worksheets[sht]
                    dst_ws = dst_wb.worksheets[sht]

                    first_data_rd_cell_in_shts_a1 = coordinate_from_string(first_data_rd_cell_in_shts[sht])

                    first_data_cell_in_sht_rc = coordinate_to_tuple(first_data_cell_in_shts[sht])

                    pg_size_src = (
                        conf[stp][rpt_type][src_rpt_cycl]["pg_head_size_in_shts"][sht]
                        + src_rpt_date_height
                        + conf[stp][rpt_type][src_rpt_cycl]["pg_tail_size_in_shts"][sht]
                    )
                    pg_size_dst = (
                        conf[stp][rpt_type][dst_rpt_cycl]["pg_head_size_in_shts"][sht]
                        + dst_rpt_date_height
                        + conf[stp][rpt_type][dst_rpt_cycl]["pg_tail_size_in_shts"][sht]
                    )

                    # monthly yearly report sheet page level
                    for pg in range(0, wb_pgs_in_shts[sht], 1):
                        # ### 특별조건 시작
                        # 경안  ->> 나중에 원본 xlsx 수정필요
                        if (conf[stp]["name"] == "경안맑은물복원센터") and (sht == 8):
                            if (pg >= 1) and (pg < 10):
                                pg_offset = 1
                            elif pg >= 10:
                                pg_offset = 2
                            else:
                                pg_offset = 0
                        # 오포 ->> 나중에 원본 xlsx 수정필요
                        elif conf[stp]["name"] == "오포맑은물복원센터":
                            if sht == 0:
                                if pg >= 1:
                                    pg_offset = 3
                            elif sht == 6:
                                if pg >= 1:
                                    pg_offset = 5
                            else:
                                pg_offset = 0

                        else:
                            pg_offset = 0
                        # ### 특별조건 끝 ->>

                        # Month, Year Date Calculation & Write
                        # Date Print
                        if rpt_idx == 1:
                            date_cell_in_shts_a1 = coordinate_from_string(date_cell_in_shts[sht])

                            pg_date_cell_a1 = (
                                f"{date_cell_in_shts_a1[0]}"
                                f"{str(date_cell_in_shts_a1[1] + ((pg_size_dst * pg) + pg_offset))}"
                            )
                            if rpt_type == YEARLY:  # yearly report
                                dst_ws[pg_date_cell_a1].value = datetime.date(start_date.year, 1, 1)
                                dst_ws[pg_date_cell_a1].number_format = "yyyy"
                            else:  # monthly
                                dst_ws[pg_date_cell_a1].value = datetime.date(start_date.year, start_date.month, 1)
                                dst_ws[pg_date_cell_a1].number_format = "yyyy-mm"
                        # Read src data
                        pg_sum_first_cell = (
                            f"{first_data_rd_cell_in_shts_a1[0]}"
                            f"{first_data_rd_cell_in_shts_a1[1] + ((pg_size_src * pg) + pg_offset)}"
                        )
                        pg_sum_last_cell_col = get_column_letter(
                            column_index_from_string(first_data_rd_cell_in_shts_a1[0]) + wb_cols_in_shts[sht][pg] - 1
                        )
                        pg_sum_last_cell = (
                            f"{pg_sum_last_cell_col}"
                            f"{first_data_rd_cell_in_shts_a1[1] + 3 + ((pg_size_src * pg) + pg_offset )}"
                        )
                        pg_sum_rng = f"{pg_sum_first_cell}:{pg_sum_last_cell}"

                        xl_sum_cell_rng = src_ws[pg_sum_rng]

                        # Write src data
                        for i in range(0, len(wb_sum_cell_types[sht][pg]), 1):
                            j = wb_sum_cell_types[sht][pg][i]
                            r1 = first_data_cell_in_sht_rc[0] + ((pg_size_dst * pg) + rpt_idx + pg_offset) - 1
                            c1 = first_data_cell_in_sht_rc[1] + i
                            # print(sht, pg, get_column_letter(c1) + str(r1), i) # <= check point
                            dst_ws[f"{get_column_letter(c1)}{r1}"].value = xl_sum_cell_rng[j][i].value

            src_wb.close()

        finally:
            pass

    save_folder = os.path.dirname(dst_wb_path)

    try:
        if not os.path.isdir(save_folder):
            os.makedirs(save_folder)
    except OSError:
        print("Error: Failed to create the directory.")

    dst_wb.save(filename=f"{dst_wb_path}/{dst_wb_name}")
    dst_wb.close()

    # First calculation formular
    # openpyxl은 엑셀 함수식을 계산하지 못하여 엑셀을 이용하여 계산식을 계산후 저장한다(년보생산시 월보데이터 계산결과 미생성 방지)
    excel = win32com.client.Dispatch("Excel.Application")
    excel.ScreenUpdating = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False
    excel.Interactive = False
    excel.Visible = False
    wb = excel.Workbooks.Open(f"{dst_wb_path}/{dst_wb_name}")
    time.sleep(5)
    wb.Save()
    wb.Close()


if __name__ == "__main__":
    import warnings

    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    # close all excel app
    xlrpt_utils.close_all_excel_app()
    # read configuration file
    xconf = xlrpt_utils.read_config()

    # 시작일,종료일 설정
    START_DATE = "2023-01-01"
    END_DATE = "2023-01-31"

    # 시작일, 종료일 datetime 으로 변환
    dt_start_date = datetime.datetime.strptime(START_DATE, "%Y-%m-%d")
    dt_last_date = xlrpt_utils.last_day_of_month(dt_start_date)
    dt_end_date = datetime.datetime.strptime(END_DATE, "%Y-%m-%d")

    xlsx_rpt(conf=xconf, stp="op", rpt_type="op", rpt_cycl="m", start_date=dt_start_date, end_date=dt_last_date)

    print("End of Test")
